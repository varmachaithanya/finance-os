import csv
import os
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO

from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.config import settings
from app.repositories.expense_repository import ExpenseRepository
from app.repositories.income_repository import IncomeRepository


class ReportService:
    def __init__(self, db):
        self.db = db
        self.expense_repo = ExpenseRepository(db)
        self.income_repo = IncomeRepository(db)

    def _get_date_range(self, period: str, year: int, month: int = None, week: int = None):
        today = date.today()
        if period == "daily":
            return today, today
        elif period == "weekly":
            if week:
                start = date(year, 1, 1) + timedelta(weeks=week - 1)
                return start, start + timedelta(days=6)
            start = today - timedelta(days=today.weekday())
            return start, start + timedelta(days=6)
        elif period == "monthly":
            if month:
                start = date(year, month, 1)
                if month == 12:
                    end = date(year + 1, 1, 1) - timedelta(days=1)
                else:
                    end = date(year, month + 1, 1) - timedelta(days=1)
                return start, end
            start = today.replace(day=1)
            if today.month == 12:
                end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
            return start, end
        elif period == "yearly":
            start = date(year, 1, 1)
            end = date(year, 12, 31)
            return start, end
        return today.replace(day=1), today

    def generate(self, user_id: str, period: str = "monthly", year: int = None, month: int = None, week: int = None) -> dict:
        today = date.today()
        year = year or today.year
        month = month or today.month
        start_date, end_date = self._get_date_range(period, year, month, week)

        total_income = self.income_repo.get_total_for_period(user_id, start_date, end_date)
        total_expenses = self.expense_repo.get_total_for_period(user_id, start_date, end_date)

        expenses, _ = self.expense_repo.get_by_user(
            user_id, skip=0, limit=10000, from_date=start_date, to_date=end_date
        )
        income, _ = self.income_repo.get_by_user(
            user_id, skip=0, limit=10000, from_date=start_date, to_date=end_date
        )

        return {
            "summary": {
                "period": period,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "total_income": total_income,
                "total_expenses": total_expenses,
                "net_savings": total_income - total_expenses,
            },
            "expenses": [
                {
                    "date": e.expense_date.isoformat(),
                    "description": e.description,
                    "amount": e.amount,
                    "payment_method": e.payment_method,
                }
                for e in expenses
            ],
            "income": [
                {
                    "date": i.income_date.isoformat(),
                    "source": i.source,
                    "amount": i.amount,
                    "description": i.description,
                }
                for i in income
            ],
            "generated_at": datetime.now().isoformat(),
        }

    def export_pdf(self, user_id: str, period: str = "monthly", year: int = None, month: int = None, week: int = None) -> str:
        data = self.generate(user_id, period, year, month, week)
        os.makedirs(settings.REPORTS_DIR, exist_ok=True)
        filename = f"report_{user_id}_{period}_{date.today().isoformat()}.pdf"
        filepath = os.path.join(settings.REPORTS_DIR, filename)

        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(f"WealthWise Report - {period.title()}", styles["Title"]))
        elements.append(Spacer(1, 12))
        elements.append(
            Paragraph(
                f"Period: {data['summary']['start_date']} to {data['summary']['end_date']}",
                styles["Normal"],
            )
        )
        elements.append(Spacer(1, 12))

        summary_data = [
            ["Metric", "Amount"],
            ["Total Income", f"₹{data['summary']['total_income']:,.2f}"],
            ["Total Expenses", f"₹{data['summary']['total_expenses']:,.2f}"],
            ["Net Savings", f"₹{data['summary']['net_savings']:,.2f}"],
        ]
        summary_table = Table(summary_data)
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1976d2")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 12),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ]
            )
        )
        elements.append(summary_table)
        elements.append(Spacer(1, 24))

        if data["expenses"]:
            elements.append(Paragraph("Expense Breakdown", styles["Heading2"]))
            expense_data = [["Date", "Description", "Amount"]]
            for e in data["expenses"]:
                expense_data.append([e["date"], e["description"] or "-", f"₹{e['amount']:,.2f}"])
            expense_table = Table(expense_data)
            expense_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ]
                )
            )
            elements.append(expense_table)

        doc.build(elements)
        return filepath

    def export_excel(self, user_id: str, period: str = "monthly", year: int = None, month: int = None, week: int = None) -> str:
        data = self.generate(user_id, period, year, month, week)
        os.makedirs(settings.REPORTS_DIR, exist_ok=True)
        filename = f"report_{user_id}_{period}_{date.today().isoformat()}.xlsx"
        filepath = os.path.join(settings.REPORTS_DIR, filename)

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")

        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Metric", "Amount"])
        ws_summary.append(["Total Income", float(data["summary"]["total_income"])])
        ws_summary.append(["Total Expenses", float(data["summary"]["total_expenses"])])
        ws_summary.append(["Net Savings", float(data["summary"]["net_savings"])])
        for cell in ws_summary[1]:
            cell.font = header_font
            cell.fill = header_fill

        ws_expenses = wb.create_sheet("Expenses")
        ws_expenses.append(["Date", "Description", "Amount"])
        for cell in ws_expenses[1]:
            cell.font = header_font
            cell.fill = header_fill
        for e in data["expenses"]:
            ws_expenses.append([e["date"], e["description"], float(e["amount"])])

        ws_income = wb.create_sheet("Income")
        ws_income.append(["Date", "Source", "Amount"])
        for cell in ws_income[1]:
            cell.font = header_font
            cell.fill = header_fill
        for i in data["income"]:
            ws_income.append([i["date"], i["source"], float(i["amount"])])

        wb.save(filepath)
        return filepath

    def export_csv(self, user_id: str, period: str = "monthly", year: int = None, month: int = None, week: int = None) -> str:
        data = self.generate(user_id, period, year, month, week)
        os.makedirs(settings.REPORTS_DIR, exist_ok=True)
        filename = f"report_{user_id}_{period}_{date.today().isoformat()}.csv"
        filepath = os.path.join(settings.REPORTS_DIR, filename)

        with open(filepath, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Type", "Date", "Description/Source", "Amount"])
            for e in data["expenses"]:
                writer.writerow(["Expense", e["date"], e["description"], float(e["amount"])])
            for i in data["income"]:
                writer.writerow(["Income", i["date"], i["source"], float(i["amount"])])

        return filepath
